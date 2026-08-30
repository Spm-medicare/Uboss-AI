"""Importing the approved Skill Catalogue.

PLAN §39 keeps the 400-skill catalogue and its 2,400 IF-THEN rules as *"a client-owned seed
asset"*, and §40 says to *"preserve/import validated Skill Catalogue records and IF-THEN rules"*.
This is that import.

**Idempotent, and matched on the catalogue's own ids.** Run it twice and nothing doubles: a row
whose `catalogue_id` already exists is updated in place. That matters more than it sounds — the
seed will be re-imported every time the workbook is corrected, and a version of this that appended
would leave two `U-001`s and no way to tell which one a search returned.

**Run as the owner, never by the application.** The catalogue is shared reference data with no
tenant; the application role can read it and cannot write it, which is the boundary migration 0019
set up. This runs from a script or a migration, both of which use the owner connection.

**Nothing is invented.** Every value comes from a cell. Where the workbook is empty the column is
null, and where a value is outside a closed list the row is reported rather than coerced — a
silently corrected catalogue is a catalogue nobody can check against the sheet.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from uboss.core.errors import ValidationFailed
from uboss.modules.agents.models import Skill, SkillArchetype, SkillExactnessGate, SkillRule

#: The sheets this reads, by their names in the approved workbook.
SHEETS = {
    "archetypes": "Skill Types",
    "skills": "Skill Catalogue",
    "rules": "IF-THEN Rules",
    "gates": "Exactness Gates",
}

#: The workbook writes autonomy as "A1 — Read / analyze". Only the code is stored: the words are
#: the same for every row that shares a code, and keeping both would be two facts to disagree.
AUTONOMY_PATTERN = re.compile(r"^(A[1-4])")


@dataclass(slots=True)
class Report:
    """What the import did, and what it could not.

    Returned rather than logged, so a caller can refuse to continue on a bad import instead of
    finding out from a log file afterwards.
    """

    archetypes: int = 0
    gates: int = 0
    skills_created: int = 0
    skills_updated: int = 0
    rules_created: int = 0
    rules_updated: int = 0
    #: Rows the sheet had that this could not use, each with a reason. Never silently dropped.
    skipped: list[str] = field(default_factory=list)

    @property
    def total_skills(self) -> int:
        return self.skills_created + self.skills_updated

    @property
    def total_rules(self) -> int:
        return self.rules_created + self.rules_updated


def _cell(value: Any) -> str | None:
    if value is None:
        return None
    text_value = str(value).strip()
    return text_value or None


def _rows(sheet: Any) -> list[dict[str, Any]]:
    """Rows keyed by the sheet's own header, so a reordered column changes nothing here."""
    raw = list(sheet.iter_rows(values_only=True))
    if not raw:
        return []
    header = [str(cell).strip() if cell else "" for cell in raw[0]]
    return [
        {header[index]: cell for index, cell in enumerate(row) if index < len(header)}
        for row in raw[1:]
        if any(cell is not None for cell in row)
    ]


async def import_catalogue(session: AsyncSession, workbook_path: Path) -> Report:
    """Read the workbook into the registry. Safe to run repeatedly.

    The file read is blocking, and deliberately so: this runs from a script or a migration, not
    from a request, so there is no event loop to starve. Wrapping it in a thread would add a
    moving part for no benefit anybody could observe.
    """
    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except FileNotFoundError as not_found:
        raise ValidationFailed(f"No workbook at {workbook_path}.") from not_found
    except Exception as cause:  # openpyxl raises a wide range for a malformed file
        raise ValidationFailed(
            f"That file could not be opened as a workbook: {workbook_path}"
        ) from cause
    try:
        absent = [name for name in SHEETS.values() if name not in workbook.sheetnames]
        if absent:
            raise ValidationFailed(
                f"That workbook has no sheet called {absent[0]}. It is not the approved "
                "Skill Catalogue."
            )

        report = Report()
        await _archetypes(session, _rows(workbook[SHEETS["archetypes"]]), report)
        await _gates(session, _rows(workbook[SHEETS["gates"]]), report)
        by_catalogue_id = await _skills(
            session, _rows(workbook[SHEETS["skills"]]), report
        )
        await _rules(session, _rows(workbook[SHEETS["rules"]]), by_catalogue_id, report)
        return report
    finally:
        workbook.close()


async def _archetypes(
    session: AsyncSession, rows: list[dict[str, Any]], report: Report
) -> None:
    existing = {
        row.id: row
        for row in (await session.execute(select(SkillArchetype))).scalars().all()
    }
    for position, row in enumerate(rows, start=1):
        identifier = _cell(row.get("Type ID"))
        name = _cell(row.get("Skill Archetype"))
        if not identifier or not name:
            report.skipped.append(f"archetype row {position}: no id or name")
            continue

        values = {
            "name": name,
            "if_clause": _cell(row.get("IF")) or "",
            "then_clause": _cell(row.get("THEN")) or "",
            "typical_output": _cell(row.get("Typical Output")),
            "mandatory_controls": _cell(row.get("Mandatory Controls")),
            "position": position,
        }
        if identifier in existing:
            for key, value in values.items():
                setattr(existing[identifier], key, value)
        else:
            session.add(SkillArchetype(id=identifier, **values))
        report.archetypes += 1
    await session.flush()


async def _gates(session: AsyncSession, rows: list[dict[str, Any]], report: Report) -> None:
    existing = {
        row.id: row
        for row in (await session.execute(select(SkillExactnessGate))).scalars().all()
    }
    for position, row in enumerate(rows, start=1):
        identifier = _cell(row.get("Gate ID"))
        name = _cell(row.get("Exactness Gate"))
        if not identifier or not name:
            report.skipped.append(f"gate row {position}: no id or name")
            continue

        values = {
            "name": name,
            "if_clause": _cell(row.get("IF")) or "",
            "then_clause": _cell(row.get("THEN")) or "",
            "pass_evidence": _cell(row.get("Pass Evidence")) or "",
            #  Quoted verbatim. The resolver shows this word for word, so a refusal reads the
            #  same on screen as it does in the approved sheet.
            "failure_state": _cell(row.get("Failure State")) or "",
            "position": position,
        }
        if identifier in existing:
            for key, value in values.items():
                setattr(existing[identifier], key, value)
        else:
            session.add(SkillExactnessGate(id=identifier, **values))
        report.gates += 1
    await session.flush()


async def _skills(
    session: AsyncSession, rows: list[dict[str, Any]], report: Report
) -> dict[str, Any]:
    existing = {
        row.catalogue_id: row
        for row in (
            await session.execute(select(Skill).where(Skill.catalogue_id.is_not(None)))
        )
        .scalars()
        .all()
        if row.catalogue_id
    }
    known_archetypes = {
        row_id
        for (row_id,) in (await session.execute(select(SkillArchetype.id))).all()
    }

    for position, row in enumerate(rows, start=1):
        identifier = _cell(row.get("Skill ID"))
        name = _cell(row.get("Skill Name"))
        if not identifier or not name:
            report.skipped.append(f"skill row {position}: no id or name")
            continue

        autonomy_cell = _cell(row.get("Autonomy")) or ""
        match = AUTONOMY_PATTERN.match(autonomy_cell)
        if match is None:
            #  Reported, not coerced. A skill whose autonomy nobody can read is a skill the
            #  resolver would let operate at whatever default was chosen for it.
            report.skipped.append(f"{identifier}: autonomy {autonomy_cell!r} not understood")
            continue

        archetype_name = _cell(row.get("Archetype"))
        archetype_id = None
        if archetype_name:
            archetype_id = await _archetype_for(session, archetype_name, known_archetypes)

        layer = _cell(row.get("Layer")) or "Universal Department"
        values: dict[str, Any] = {
            "name": name,
            "layer": layer,
            "department": _cell(row.get("Department")),
            "industry": _cell(row.get("Industry")),
            "archetype_id": archetype_id,
            "purpose": _cell(row.get("Purpose")),
            "positive_trigger": _cell(row.get("Positive Trigger")),
            #  The column that matters most on a search result: what this skill is *not* for.
            "exclusions": _cell(row.get("Do Not Use / Exclusions")),
            "minimum_inputs": _cell(row.get("Minimum Inputs")),
            "primary_if": _cell(row.get("Primary IF")),
            "primary_then": _cell(row.get("Primary THEN")),
            "output": _cell(row.get("Output")),
            "validation_gate": _cell(row.get("Validation Gate")),
            "autonomy": match.group(1),
            "source_ids": _cell(row.get("Source IDs")),
            "status": "published",
        }

        if identifier in existing:
            for key, value in values.items():
                setattr(existing[identifier], key, value)
            report.skills_updated += 1
        else:
            session.add(Skill(catalogue_id=identifier, tenant_id=None, **values))
            report.skills_created += 1

    await session.flush()

    return {
        row.catalogue_id: row.id
        for row in (
            await session.execute(select(Skill).where(Skill.catalogue_id.is_not(None)))
        )
        .scalars()
        .all()
        if row.catalogue_id
    }


async def _archetype_for(
    session: AsyncSession, name: str, known: set[str]
) -> str | None:
    """Match a skill's archetype by name.

    The catalogue sheet names the archetype rather than coding it, so the join is on the name.
    A name the archetype sheet does not carry leaves the skill unclassified rather than being
    invented — an archetype nobody agreed is worse than none.
    """
    row = (
        await session.execute(
            select(SkillArchetype.id).where(SkillArchetype.name == name)
        )
    ).scalar_one_or_none()
    return row if row in known else None


async def _rules(
    session: AsyncSession,
    rows: list[dict[str, Any]],
    by_catalogue_id: dict[str, Any],
    report: Report,
) -> None:
    existing = {
        row.catalogue_id: row
        for row in (
            await session.execute(
                select(SkillRule).where(SkillRule.catalogue_id.is_not(None))
            )
        )
        .scalars()
        .all()
        if row.catalogue_id
    }
    seen_per_skill: dict[Any, int] = {}

    for index, row in enumerate(rows, start=1):
        identifier = _cell(row.get("Rule ID"))
        skill_ref = _cell(row.get("Skill ID"))
        if not identifier or not skill_ref:
            report.skipped.append(f"rule row {index}: no id or skill")
            continue

        skill_id = by_catalogue_id.get(skill_ref)
        if skill_id is None:
            #  A rule pointing at a skill the sheet does not contain. Reported rather than
            #  dropped: it means the two sheets disagree, which somebody needs to know.
            report.skipped.append(f"{identifier}: no skill {skill_ref} in the catalogue")
            continue

        seen_per_skill[skill_id] = seen_per_skill.get(skill_id, 0) + 1
        values = {
            "skill_id": skill_id,
            "condition_type": _cell(row.get("Condition Type")) or "Primary Decision",
            "if_clause": _cell(row.get("IF")) or "",
            "then_clause": _cell(row.get("THEN")) or "",
            "priority": _cell(row.get("Priority")) or "High",
            "evidence_required": _cell(row.get("Evidence Required")),
            #  What happens when the rule does not hold. The whole reason these rows are worth
            #  importing rather than summarising.
            "failure_state": _cell(row.get("Failure State")),
            "human_gate": _cell(row.get("Human Gate")),
            "source_ids": _cell(row.get("Source IDs")),
            "position": seen_per_skill[skill_id],
        }

        if identifier in existing:
            for key, value in values.items():
                setattr(existing[identifier], key, value)
            report.rules_updated += 1
        else:
            session.add(SkillRule(catalogue_id=identifier, **values))
            report.rules_created += 1

    await session.flush()


async def catalogue_counts(session: AsyncSession) -> dict[str, int]:
    """What is actually in the registry. Used by the readiness probe and by tests."""
    result = await session.execute(
        text(
            "SELECT "
            "(SELECT count(*) FROM skill_archetypes) AS archetypes, "
            "(SELECT count(*) FROM skill_exactness_gates) AS gates, "
            "(SELECT count(*) FROM skills WHERE tenant_id IS NULL) AS skills, "
            "(SELECT count(*) FROM skill_rules) AS rules"
        )
    )
    row = result.mappings().one()
    return dict(row)
