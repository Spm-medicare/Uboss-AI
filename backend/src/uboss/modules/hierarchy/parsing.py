"""Reading a spreadsheet, deterministically.

PLAN §5 step 2: *"Parse sheets and columns deterministically."* Everything in this module is pure
and has no model in it. That ordering is the point — a model is asked only about the columns this
code could not decide, and there are usually none.

**The file is untrusted input.** It came from a customer's HR system by way of somebody's laptop.
So: read-only workbook, values only (no formula is evaluated, no linked workbook is fetched), a
row cap, and a cell-length cap. A spreadsheet parser is a common way into a server, and "it is
only an org chart" is what everybody says beforehand.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from typing import Any

from openpyxl import load_workbook

from uboss.core.errors import ValidationFailed

#: Rows past this are refused rather than truncated. An organisation with more than this many
#: rows in one file is doing something the reviewer cannot meaningfully check anyway, and
#: silently dropping the tail would produce a tree missing a division nobody noticed.
MAX_ROWS = 20_000

#: Long enough for any real name, short enough that a 2 MB cell cannot become 20,000 of them.
MAX_CELL_LENGTH = 500

#: More than this and the sheet is not an org chart.
MAX_COLUMNS = 100


@dataclass(frozen=True, slots=True)
class Sheet:
    """One parsed sheet: its header and its rows, as strings."""

    name: str | None
    columns: list[str]
    #: Row number as the spreadsheet shows it (header is 1, so data starts at 2), and the cells
    #: keyed by column name. Kept as strings — interpreting them is mapping's job, not parsing's.
    rows: list[tuple[int, dict[str, str]]]


# ------------------------------------------------------------------- the fields we know


@dataclass(frozen=True, slots=True)
class Field:
    """A column the product understands, and the header spellings that mean it.

    `aliases` is deliberately a plain list of normalised strings rather than a fuzzy matcher.
    Fuzzy matching is how "Manager Email" quietly becomes "Manager", and the failure is silent.
    A header that is not on this list is *ambiguous*, which is a state with a person in it.
    """

    name: str
    aliases: tuple[str, ...]
    required: bool = False
    description: str = ""


FIELDS: tuple[Field, ...] = (
    Field(
        "unit_name",
        ("department", "departmentname", "unit", "unitname", "team", "orgunit", "division"),
        required=True,
        description="The department, team or division this row describes.",
    ),
    Field(
        "parent_name",
        ("parent", "parentdepartment", "parentunit", "reportsto", "belongsto"),
        description="The department this one sits inside. Blank means the top of the tree.",
    ),
    Field(
        "unit_type",
        ("type", "unittype", "level", "orglevel"),
        description="company, division, department or team.",
    ),
    Field(
        "unit_ref",
        ("code", "costcentre", "costcenter", "departmentcode", "unitcode", "externalid"),
        description="The customer's own identifier for the department.",
    ),
    Field(
        "position_title",
        ("position", "positiontitle", "jobtitle", "role", "designation"),
        description="The seat. A row with one of these creates a position in the department.",
    ),
    Field(
        "position_ref",
        ("positioncode", "positionid", "seatid"),
        description="The customer's own identifier for the position.",
    ),
    Field(
        "location",
        ("location", "site", "office", "city"),
        description="Where the department or position sits.",
    ),
    Field(
        "person_email",
        ("email", "emailaddress", "workemail", "personemail"),
        description="Who holds the position. Matched against people already in the workspace.",
    ),
    Field(
        "person_name",
        ("name", "fullname", "employeename", "personname"),
        description="Shown while reviewing. Never used to match a person — the address is.",
    ),
    Field(
        "effective_from",
        ("effectivefrom", "startdate", "from", "joineddate"),
        description="The date the assignment starts. Defaults to today.",
    ),
)

FIELDS_BY_NAME: dict[str, Field] = {item.name: item for item in FIELDS}


def normalise(header: str) -> str:
    """`"Cost Centre "` and `"cost_centre"` are the same header. `"Cost Centre 2"` is not."""
    return re.sub(r"[^a-z0-9]", "", header.strip().lower())


def match_columns(columns: list[str]) -> tuple[dict[str, str], list[str]]:
    """Split the header into what is understood and what is not.

    Returns `(mapping, ambiguous)` where `mapping` is `{column: field}` for the exact matches and
    `ambiguous` lists the headers nothing matched.

    A second column claiming a field the first already took is left ambiguous rather than
    silently overwriting. Two columns both called "Code" is a real spreadsheet, and guessing
    which one was meant is exactly the guess that produces a wrong tree.
    """
    mapping: dict[str, str] = {}
    taken: set[str] = set()
    ambiguous: list[str] = []

    for column in columns:
        key = normalise(column)
        matched = next(
            (item.name for item in FIELDS if key in item.aliases or key == item.name),
            None,
        )
        if matched is None or matched in taken:
            ambiguous.append(column)
            continue
        mapping[column] = matched
        taken.add(matched)

    return mapping, ambiguous


# ------------------------------------------------------------------------------ reading


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:MAX_CELL_LENGTH]


def _finish(name: str | None, header: list[str], body: list[list[str]]) -> Sheet:
    columns = [_clean(cell) for cell in header]
    if not any(columns):
        raise ValidationFailed("That file has no header row, so its columns have no names.")
    if len(columns) > MAX_COLUMNS:
        raise ValidationFailed(f"That sheet has more than {MAX_COLUMNS} columns.")

    #  Blank headers become positional names rather than being dropped: dropping one would shift
    #  every cell after it into the wrong column, which is the worst possible silent failure.
    columns = [column or f"Column {index + 1}" for index, column in enumerate(columns)]

    rows: list[tuple[int, dict[str, str]]] = []
    for offset, raw in enumerate(body):
        cells = {
            column: _clean(raw[index]) if index < len(raw) else ""
            for index, column in enumerate(columns)
        }
        #  A wholly blank line is a spreadsheet artefact, not a row somebody meant.
        if not any(cells.values()):
            continue
        rows.append((offset + 2, cells))

    return Sheet(name=name, columns=columns, rows=rows)


def read_csv(data: bytes) -> Sheet:
    """Read a CSV.

    The delimiter is sniffed from the first few kilobytes. A file that is not delimited at all
    falls back to commas rather than being refused — one column is a valid, if unhelpful, sheet,
    and the mapping step will say so more clearly than a parser error could.
    """
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        #  Excel on a Windows machine still writes this, and refusing it would mean refusing a
        #  large fraction of real files for no reason a customer could act on.
        try:
            text = data.decode("cp1252")
        except UnicodeDecodeError as cause:
            raise ValidationFailed(
                "That file is not text this reader understands. Save it as UTF-8 CSV or .xlsx."
            ) from cause

    sample = text[:4096]
    try:
        dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(sample, ",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header = next(reader)
    except StopIteration:
        raise ValidationFailed("That file is empty.") from None

    body: list[list[str]] = []
    for row in reader:
        if len(body) >= MAX_ROWS:
            raise ValidationFailed(
                f"That file has more than {MAX_ROWS:,} rows. Split it and import in parts."
            )
        body.append(row)

    return _finish(None, header, body)


def read_xlsx(data: bytes, sheet_name: str | None = None) -> Sheet:
    """Read one sheet of a workbook.

    `read_only` and `data_only`: no formula is evaluated and no linked workbook is fetched. Both
    matter — this file arrived from outside, and a spreadsheet that can reach the network from
    inside the server is a spreadsheet that can reach anything the server can.
    """
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_links=False)
    except Exception as cause:  # openpyxl raises a wide range for a malformed file
        raise ValidationFailed(
            "That .xlsx could not be opened. It may be a different format, or damaged."
        ) from cause

    try:
        if sheet_name is not None:
            if sheet_name not in workbook.sheetnames:
                raise ValidationFailed(f"That workbook has no sheet called “{sheet_name}”.")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows = worksheet.iter_rows(values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            raise ValidationFailed("That sheet is empty.") from None

        body: list[list[str]] = []
        for row in rows:
            if len(body) >= MAX_ROWS:
                raise ValidationFailed(
                    f"That sheet has more than {MAX_ROWS:,} rows. Split it and import in parts."
                )
            body.append([_clean(cell) for cell in row])

        return _finish(worksheet.title, [_clean(cell) for cell in header], body)
    finally:
        workbook.close()


def sheet_names(data: bytes) -> list[str]:
    """What sheets a workbook has, so the person can pick one before anything is parsed."""
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, keep_links=False)
    except Exception as cause:
        raise ValidationFailed("That .xlsx could not be opened.") from cause
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read(data: bytes, filename: str, sheet_name: str | None = None) -> Sheet:
    """Pick a reader from the file's name.

    The extension, not the browser's content type: a browser sends whatever the operating system
    associated with the extension, and that is frequently wrong for `.csv`.
    """
    lowered = filename.lower()
    if lowered.endswith(".csv") or lowered.endswith(".txt"):
        return read_csv(data)
    if lowered.endswith(".xlsx") or lowered.endswith(".xlsm"):
        return read_xlsx(data, sheet_name)
    raise ValidationFailed(
        "Import a .csv or .xlsx file. The older .xls format is not supported — open it in a "
        "spreadsheet application and save it as .xlsx."
    )


# ----------------------------------------------------------------------------- staging


@dataclass(slots=True)
class StagedRow:
    """One spreadsheet row, understood."""

    row_number: int
    raw: dict[str, str]
    kind: str
    parsed: dict[str, Any]
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


UNIT_TYPES = ("company", "division", "department", "team")


def stage(sheet: Sheet, mapping: dict[str, str]) -> list[StagedRow]:
    """Turn mapped rows into what they mean, and say what is wrong with each.

    Nothing here touches the database. That is what makes the preview trustworthy: the same
    function produced what the person is looking at and what will be applied.

    Errors stop a row; warnings do not. The distinction is whether the product could apply the
    row and be right about it — a missing department name means there is nothing to create, while
    an unrecognised unit type means the product picked `department` and should say so.
    """
    fields = set(mapping.values())
    if "unit_name" not in fields:
        raise ValidationFailed(
            "No column was mapped to the department name, so there is nothing to build a tree "
            "from. Map one before continuing."
        )

    by_field = {value: key for key, value in mapping.items()}
    staged: list[StagedRow] = []
    seen_units: dict[str, int] = {}
    #: code -> (the department that claimed it, the row it was first seen on). A spreadsheet
    #: lists one department once per position, so the same code arriving again for the same
    #: department is normal. The same code on a *different* department is the duplicate
    #: identifier PLAN §5 asks the product to catch.
    #: Keyed on the lowercased code; the department is kept as it was written, because the
    #: message names it back to the person and "one" is not what they typed.
    seen_refs: dict[str, tuple[str, str, int]] = {}

    for row_number, cells in sheet.rows:
        #  Bound as a default rather than captured: a closure over the loop variable would read
        #  the last row's cells for every row, and the result would look plausible.
        def value(name: str, cells: dict[str, str] = cells) -> str:
            column = by_field.get(name)
            return cells.get(column, "") if column else ""

        unit_name = value("unit_name")
        title = value("position_title")
        errors: list[str] = []
        warnings: list[str] = []

        if not unit_name:
            errors.append("No department name in this row.")

        unit_type = value("unit_type").lower()
        if unit_type and unit_type not in UNIT_TYPES:
            warnings.append(
                f"“{unit_type}” is not a known type, so this will be created as a department."
            )
            unit_type = ""

        unit_ref = value("unit_ref")
        if unit_ref:
            claimed_key, claimed_name, first_row = seen_refs.get(
                unit_ref.lower(), ("", "", 0)
            )
            if claimed_key and claimed_key != unit_name.lower():
                errors.append(
                    f"Code “{unit_ref}” is used by “{claimed_name}” on row {first_row} "
                    f"and by “{unit_name}” here."
                )
            elif not claimed_key:
                seen_refs[unit_ref.lower()] = (unit_name.lower(), unit_name, row_number)

        #  Two rows naming the same department is normal — that is how a spreadsheet lists two
        #  positions in one team. It is only a problem if they disagree about its parent.
        key = unit_name.lower()
        parent = value("parent_name")
        if key in seen_units and parent:
            previous = staged[seen_units[key]]
            earlier_parent = str(previous.parsed.get("parent_name") or "")
            if earlier_parent and earlier_parent.lower() != parent.lower():
                errors.append(
                    f"“{unit_name}” is under “{earlier_parent}” on row "
                    f"{previous.row_number} and under “{parent}” here."
                )
        elif key not in seen_units:
            seen_units[key] = len(staged)

        email = value("person_email")
        if email and not title:
            warnings.append("Somebody is named here but there is no position for them to hold.")
        if email and "@" not in email:
            errors.append(f"“{email}” is not an email address.")

        parsed: dict[str, Any] = {
            "unit_name": unit_name,
            "parent_name": parent,
            "unit_type": unit_type or "department",
            "unit_ref": unit_ref,
            "position_title": title,
            "position_ref": value("position_ref"),
            "location": value("location"),
            "person_email": email,
            "person_name": value("person_name"),
            "effective_from": value("effective_from"),
        }

        #  A row with errors is still staged and still shown. A row nobody can see is a row
        #  nobody can fix, and "12 rows failed" with no list of which is a dead end.
        staged.append(
            StagedRow(
                row_number=row_number,
                raw=cells,
                kind="position" if title else "org_unit",
                parsed=parsed,
                errors=errors,
                warnings=warnings,
            )
        )

    return staged
